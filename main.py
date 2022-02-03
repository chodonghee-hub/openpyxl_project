from openpyxl import Workbook
from openpyxl import load_workbook

def SaveData(wb):
    wb.save("C:/Users/cho donghee/Desktop/som/test_cel.xlsx")

def UpdateDataToCell():
    wb = Workbook()

    ws = wb.create_sheet('생성시트')

    ws = wb.active
    ws['A1'] = '숫자'

    ws.append([1,2,3])

    ws.cell(5,5, '5행 5열')
    SaveData(wb)

def ReadDataFromCell():
    load_wb = load_workbook("C:/Users/cho donghee/Desktop/som/test_cel.xlsx")
    load_ws = load_wb['Sheet']

    print("B2   : {}".format(load_ws['B2'].value))

    print("cell (3, 2)  : {} ".format(load_ws.cell(3,2).value))

    get_cells = load_ws['B3' : 'B6']
    print("load data in [B3 : B6]   : ", end = ' ')

    try :
        for row in get_cells:
            for cell in row :
                print(cell.vaLue)
    except :
        print("no data in cell")


    # print all the data in row
    for row in load_ws.rows :
        print(row)

    # print all the data in col
    for column in load_ws.columns:
        print(column)

    # print all the data
    all_values = []
    for row in load_ws.rows:
        row_value = []
        for cell in row :
            row_value.append(cell.value)
        all_values.append(row_value)
    print(all_values)

    load_ws.cell(3, 3, 51470)
    load_ws.cell(4, 3, 21470)
    load_ws.cell(5, 3, 1470)
    load_ws.cell(6, 3, 6470)
    SaveData(load_wb)
    # SaveData(load_wb)

if __name__ == '__main__':
    print()
    ReadDataFromCell()
    #UpdateDataToCell()



